import csv
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Color, PatternFill
import json
from io import BytesIO
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas
from PyPDF2 import PdfFileMerger, PdfFileReader, PdfFileWriter
import yaml
import os
import subprocess
import ntplib
from datetime import datetime
import pytz
import socket
from pysnmp.hlapi import *
import re
import nmap
import ipaddress



class CSVWriter:
    def __init__(self, filename:str, header:list, data:list):
        self.filename = filename
        self.header = header
        self.data = data

    def write(self):
        with open(self.filename, 'w', encoding='UTF8', newline='') as f:
            writer = csv.writer(f)

            # write the header
            writer.writerow(self.header)

            # write the data
            for row in self.data:
                writer.writerow(row)



class XLSWriter:

    '''
        Named colors for use in styles."""
        BLACK = 'FF000000'
        WHITE = 'FFFFFFFF'
        RED = 'FFFF0000'
        DARKRED = 'FF800000'
        BLUE = 'FF0000FF'
        DARKBLUE = 'FF000080'
        GREEN = 'FF00FF00'
        DARKGREEN = 'FF008000'
        YELLOW = 'FFFFFF00'
        DARKYELLOW = 'FF808000'
    '''

    def __init__(self, filename:str):
        self.filename = filename
        if not os.path.exists(self.filename):
            self.wb = Workbook()
            self.wb.save(self.filename)
        else:
            self.wb = load_workbook(self.filename)

    def add_sheet(self, sheetname:str):
        sheet = self.wb.create_sheet(sheetname)
        return sheet

    def write_bulk_data(self, sheet, data:list):
        for row in data:
            sheet.append(row)

    def write_single_row_data(self, sheetname, data, font_color=None, bold=False, italic=False):
        sheet = self.wb[sheetname]
        row = sheet.max_row + 1
        for col, value in enumerate(data, start=1):
            cell = sheet.cell(row=row, column=col, value=value)
            if font_color:
                font = Font(color=font_color, bold=bold, italic=italic)
                cell.font = font
            else:
                cell.font = Font(bold=bold, italic=italic)

    def save(self):
        self.wb.save(self.filename)


class PDFCreator:

    '''
    This PDFCreator class has the following methods:

    __init__(self, filename): Initializes the class with the filename argument, which specifies the name of the output PDF file.
    add_page(self, page): Adds a new page to the PDF. The page argument is a dictionary with the following keys:
    x: The x-coordinate of the text.
    y: The y-coordinate of the text.
    text: The text to add to the page.
    create_pdf(self): Creates the PDF file using the filename specified in the constructor and the pages added with add_page. The PDF file is created using the Reportlab library to generate the PDF content and the PyPDF2 library to merge the pages and write the output file.
    '''
    def __init__(self, filename):
        self.filename = filename
        self.pages = []
    
    def add_page(self, page):
        self.pages.append(page)
    
    def create_pdf(self):
        packet = BytesIO()
        # create a new PDF with Reportlab
        can = canvas.Canvas(packet, pagesize=letter)
        for page in self.pages:
            can.drawString(page['x'], page['y'], page['text'])
        can.save()

        # move to the beginning of the StringIO buffer
        packet.seek(0)

        # create a new PDF with PyPDF2
        new_pdf = PdfFileReader(packet)
        output = PdfFileWriter()

        # add the pages of the new PDF to the output PDF
        for i in range(new_pdf.getNumPages()):
            output.addPage(new_pdf.getPage(i))

        # write the output PDF to a file
        with open(self.filename, 'wb') as f:
            output.write(f)

class JsonToYamlConverter:
    def __init__(self, json_data):
        self.json_data = json_data
    
    def convert(self):
        try:
            yaml_data = yaml.dump(json.loads(self.json_data), default_flow_style=False)
            return yaml_data
        except ValueError as e:
            print("Error converting JSON to YAML:", e)

class JsonToCsvConverter:
    def __init__(self, json_data):
        self.json_data = json_data
        
    def convert(self, csv_file_path):
        try:
            # Load the JSON data
            data = json.loads(self.json_data)
            
            # Get the fieldnames from the first row of the JSON data
            fieldnames = list(data[0].keys())
            
            # Write the data to the CSV file
            with open(csv_file_path, 'w', newline='') as csv_file:
                writer = csv.DictWriter(csv_file, fieldnames=fieldnames)
                writer.writeheader()
                for row in data:
                    writer.writerow(row)
                    
            print(f'Successfully converted JSON to CSV and saved to {csv_file_path}')
            
        except (ValueError, KeyError) as e:
            print("Error converting JSON to CSV:", e)


class CsvToJsonConverter:
    def __init__(self, csv_file_path):
        self.csv_file_path = csv_file_path
        
    def convert(self):
        with open(self.csv_file_path, 'r') as csv_file:
            csv_reader = csv.DictReader(csv_file)
            json_data = json.dumps([row for row in csv_reader])
            return json_data
        


class PingDevice:
    def __init__(self, host):
        self.host = host

    def ping(self):
        try:
            output = subprocess.check_output(["ping", "-c", "1", self.host])
            return True
        except subprocess.CalledProcessError:
            return False




class NTPClient:
    import ntplib
    from datetime import datetime
    import pytz
  
    """
        Gets the current time from the NTP server and converts it to the specified timezone.

        Args:
            timezone (str): The timezone to convert the time to. Default is 'UTC'.

        Returns:
            dict: A dictionary with the current time in the specified timezone, in the following format:
                  {
                      'year': int,
                      'month': int,
                      'day': int,
                      'hour': int,
                      'minute': int,
                      'second': int,
                      'microsecond': int,
                      'timezone': str,
                      'offset': str
                  }
    """
    
    def __init__(self, server='pool.ntp.org'):
        self.server = server
        
    def get_time(self, timezone='CET'):
        ntp_client = ntplib.NTPClient()
        response = ntp_client.request(self.server)
        utc_time = datetime.utcfromtimestamp(response.tx_time)
        timezone_obj = pytz.timezone(timezone)
        local_time = timezone_obj.localize(utc_time)
        return {
            'year': local_time.year,
            'month': local_time.month,
            'day': local_time.day,
            'hour': local_time.hour,
            'minute': local_time.minute,
            'second': local_time.second,
            'microsecond': local_time.microsecond,
            'timezone': timezone_obj.zone,
            'offset': local_time.strftime('%z')
        }
        


class DNSLookup:
    def __init__(self, domain):
        
        """
        Creates a new DNSLookup object for performing DNS lookups.

        Args:
            domain (str): The domain name to look up.
        """
        
        self.domain = domain
        
    def get_ip(self):
        """
        Performs a DNS lookup and returns the IP address(es) associated with the domain.

        Returns:
            list: A list of IP addresses associated with the domain.
        Raises:
            socket.gaierror: If the DNS lookup fails.
        """
        ip_addresses = []
        try:
            # Resolve the domain to one or more IP addresses
            ip_addresses = socket.getaddrinfo(self.domain, None, socket.AF_INET, socket.SOCK_STREAM, socket.IPPROTO_TCP)
            ip_addresses = [addr[4][0] for addr in ip_addresses]
        except socket.gaierror as e:
            print(f"DNS lookup failed: {e}")
        return ip_addresses
    

class SNMPQuery:
    
    '''
    SNMPv2
    snmp = SNMPQuery(ip='192.168.178.1', community='MyCommunity')
    snmp.get_vendor()
    'Cisco'
    >>> 
    
    SNMPv3
    snmp = SNMPQuery(ip='NEXCOURX39', user='V3User',  auth_key='XXXXXXX', priv_key='XXXXXXX')
    snmp.get_vendor()
    'Cisco'
    
    '''
    
    def __init__(self, ip, community=None, user=None, auth_key=None, priv_key=None):
        self.ip = ip
        self.community = community
        self.user = user
        self.auth_key = auth_key
        self.priv_key = priv_key

    def get_vendor(self):
        if self.community:
            snmp_query = getCmd(
                SnmpEngine(),
                CommunityData(self.community),
                UdpTransportTarget((self.ip, 161)),
                ContextData(),
                ObjectType(ObjectIdentity('SNMPv2-MIB', 'sysDescr', 0))
            )
        else:
            snmp_query = getCmd(
                SnmpEngine(),
                UsmUserData(self.user, self.auth_key, self.priv_key, authProtocol=usmHMACSHAAuthProtocol, privProtocol=usmAesCfb128Protocol),
                UdpTransportTarget((self.ip, 161)),
                ContextData(),
                ObjectType(ObjectIdentity('SNMPv2-MIB', 'sysDescr', 0))
            )

        errorIndication, errorStatus, errorIndex, varBinds = next(snmp_query)

        if errorIndication:
            print(errorIndication)
        elif errorStatus:
            print(f"Error: {errorStatus}")
        else:
            response = str(varBinds[0])
            vendor = re.search(r'Cisco|Juniper', response, re.IGNORECASE)

            if vendor:
                return vendor.group()
            else:
                return "unknown"



class NmapScanner:
    def __init__(self, target):
        self.target = target
        self.nm = nmap.PortScanner()

    def scan(self):
        # Use Nmap to scan the target host and identify its operating system
        self.nm.scan(self.target, arguments='-O')

        # Check if the scan results contain information about the operating system
        if 'osclass' in self.nm[self.target]:
            osclass = self.nm[self.target]['osclass'][0]
            os_name = osclass['osfamily']
            os_accuracy = osclass['accuracy']
            print(f"OS: {os_name} ({os_accuracy}% accuracy)")
        else:
            print("Could not identify remote OS")


class IPv4Subnet:
    def __init__(self, subnet):
        self.subnet = ipaddress.IPv4Network(subnet)

    def get_network_id(self):
        return str(self.subnet.network_address)

    def get_first_usable_ip(self):
        return str(self.subnet.network_address + 1)

    def get_last_usable_ip(self):
        return str(self.subnet.broadcast_address - 1)

    def get_broadcast_ip(self):
        return str(self.subnet.broadcast_address)
